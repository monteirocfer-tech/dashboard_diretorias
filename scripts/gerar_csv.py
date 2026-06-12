"""
Gera o CSV consolidado (formato long) a partir do Excel de Capacitações.

Uso:
    python gerar_csv.py "Capacitações_Staff__Consolidação.xlsx" ../public/base_diretorias.csv

Lógica:
- Lê todas as abas que seguem o padrão "G. Projeto '26_<Diretoria>"
- A Diretoria é extraída do nome da aba (parte após o "_")
- Cada turma preenchida (Mes T1..T5 / Status T1..T5 / etc) gera uma linha
- Linhas sem nome de treinamento ou sem mês/status são ignoradas
"""

import sys
import datetime
import openpyxl
import csv

COLS = {
    'prioridade': 1, 'unidade': 2, 'responsavel': 3, 'diretoria_col': 4,
    'temaMacro': 5, 'nome': 6, 'acompProjeto': 7, 'contratacao': 8,
    'treinamentoStatus': 9, 'fornecedor': 10, 'modalidade': 11,
    'tipo': 12, 'horas': 13, 'investimento': 14, 'publicoAlvo': 15,
    'participantes': 16
}
TURMA_COL_START = 17
N_TURMAS = 5
TURMA_WIDTH = 7  # Mes, Data, Status, Convidados, Presentes, NPS, Justificativa

OUTPUT_HEADER = [
    'Diretoria', 'Prioridade', 'Unidade', 'Responsavel', 'TemaMacro', 'Nome',
    'Fornecedor', 'Modalidade', 'Tipo', 'Horas', 'Investimento', 'PublicoAlvo',
    'Participantes', 'Turma', 'Mes', 'Data', 'Status', 'Convidados',
    'Presentes', 'NPS', 'Justificativa'
]


def cell_value(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d')
    if v is None:
        return ''
    return v


def main(xlsx_path, out_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    abas = [s for s in wb.sheetnames if "G. Projeto '26_" in s]

    rows_out = []
    for sh in abas:
        diretoria = sh.split('_', 1)[1]
        ws = wb[sh]
        for row in range(2, ws.max_row + 1):
            nome = cell_value(ws, row, COLS['nome'])
            if not nome:
                continue

            base = {
                'Diretoria': diretoria,
                'Prioridade': cell_value(ws, row, COLS['prioridade']),
                'Unidade': cell_value(ws, row, COLS['unidade']),
                'Responsavel': cell_value(ws, row, COLS['responsavel']),
                'TemaMacro': cell_value(ws, row, COLS['temaMacro']),
                'Nome': nome,
                'Fornecedor': cell_value(ws, row, COLS['fornecedor']),
                'Modalidade': cell_value(ws, row, COLS['modalidade']),
                'Tipo': cell_value(ws, row, COLS['tipo']),
                'Horas': cell_value(ws, row, COLS['horas']),
                'Investimento': cell_value(ws, row, COLS['investimento']),
                'PublicoAlvo': cell_value(ws, row, COLS['publicoAlvo']),
                'Participantes': cell_value(ws, row, COLS['participantes']),
            }

            for t in range(N_TURMAS):
                offset = TURMA_COL_START + t * TURMA_WIDTH
                mes = cell_value(ws, row, offset)
                data = cell_value(ws, row, offset + 1)
                status = cell_value(ws, row, offset + 2)
                convidados = cell_value(ws, row, offset + 3)
                presentes = cell_value(ws, row, offset + 4)
                nps = cell_value(ws, row, offset + 5)
                justificativa = cell_value(ws, row, offset + 6)

                if mes or status or data:
                    rec = dict(base)
                    rec.update({
                        'Turma': f'T{t + 1}',
                        'Mes': mes,
                        'Data': data,
                        'Status': status,
                        'Convidados': convidados,
                        'Presentes': presentes,
                        'NPS': nps,
                        'Justificativa': justificativa,
                    })
                    rows_out.append(rec)

    with open(out_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_HEADER, delimiter=';')
        writer.writeheader()
        for r in rows_out:
            writer.writerow(r)

    print(f'Gerado {out_path} com {len(rows_out)} linhas (turmas) de {len(abas)} diretorias.')


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print('Uso: python gerar_csv.py <arquivo.xlsx> <saida.csv>')
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
